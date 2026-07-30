from __future__ import annotations

from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter


MECHANIC_COLUMNS: Tuple[Tuple[str, str], ...] = (
    ("waterOutliers", "放水未集中"),
    ("p15AvoidableDeaths", "转阶段死亡"),
    ("passageCliffMistakes", "过场失误"),
    ("p1SilverArrowMissedFights", "P1银锋射怪失误"),
    ("p1SilverArrowDeaths", "P1银锋高伤致死"),
    ("missedShadows", "P2拉弓未中幻影"),
    ("collapsingVoidSnapAiming", "崩裂甩狙"),
    ("gravityLineViolation", "重力坍缩致死违规"),
    ("voidGraspHealingLow", "空虚之握治疗不足"),
    ("tankRiftSlashFailure", "裂隙换坦失误"),
    ("voreluthVulnerabilityFade", "P1龌勒易伤"),
)

HEADERS: List[str] = (
    ["ID", "职责", "判定次数"]
    + [label for _, label in MECHANIC_COLUMNS]
    + ["申诉次数", "原因", "追加次数", "追加原因", "总计"]
)

COL_RECOGNITION = 3
COL_APPEAL = 3 + len(MECHANIC_COLUMNS) + 1
COL_APPEAL_REASON = COL_APPEAL + 1
COL_ADDITIONAL = COL_APPEAL + 2
COL_ADDITIONAL_REASON = COL_ADDITIONAL + 1
COL_TOTAL = COL_ADDITIONAL + 2
REASON_COLS = {COL_APPEAL_REASON, COL_ADDITIONAL_REASON}

TEMPLATE_PATH = Path(__file__).resolve().parent / "templates" / "verdict_excel_style.xlsx"

# Explicit styles (do NOT rely on template cell xf indices — openpyxl + template
# reuse was saving header cells as s="1" with no fill, making the header bar vanish).
HEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="FFF9FAFB")
HEADER_FILL = PatternFill("solid", fgColor="FF1F2937")
DATA_FONT = Font(name="Microsoft YaHei", size=10, color="FF111827")
DATA_FILL = PatternFill("solid", fgColor="FFF3F4F6")
TOTAL_FONT = Font(name="Microsoft YaHei", size=10, bold=True, color="FF9D174D")
TOTAL_FILL = PatternFill("solid", fgColor="FFFCE7F3")
THIN_SIDE = Side(style="thin", color="FF000000")
CELL_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

COLUMN_WIDTHS = {
    "A": 16.0,
    "B": 14.0,
    "C": 10.0,
    **{get_column_letter(4 + i): 13.0 for i in range(len(MECHANIC_COLUMNS))},
    get_column_letter(COL_APPEAL): 10.0,
    get_column_letter(COL_APPEAL_REASON): 36.0,
    get_column_letter(COL_ADDITIONAL): 10.0,
    get_column_letter(COL_ADDITIONAL_REASON): 36.0,
    get_column_letter(COL_TOTAL): 10.0,
}


def _safe_int(x: Any, default: int = 0) -> int:
    try:
        if x is None or x == "":
            return default
        return int(x)
    except Exception:
        return default


def resolve_export_dir(preferred: str | Path | None = None) -> Path:
    """优先使用可写目录：显式路径 > 桌面 > 项目 verdicts/。"""
    candidates: List[Path] = []
    if preferred:
        candidates.append(Path(preferred))
    candidates.append(Path.home() / "Desktop")
    candidates.append(Path(__file__).resolve().parents[1] / "verdicts")
    for path in candidates:
        try:
            path.mkdir(parents=True, exist_ok=True)
            probe = path / ".write_probe"
            probe.write_text("ok", encoding="utf-8")
            probe.unlink(missing_ok=True)
            return path
        except Exception:
            continue
    fallback = Path.cwd() / "verdicts"
    fallback.mkdir(parents=True, exist_ok=True)
    return fallback


def _next_available_path(path: Path) -> Path:
    """若文件存在，按 (2)/(3)... 递增追加到文件名后。"""
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    parent = path.parent
    i = 2
    while True:
        candidate = parent / f"{stem}({i}){suffix}"
        if not candidate.exists():
            return candidate
        i += 1


def _ensure_named_styles(wb: Workbook) -> None:
    """Register named styles so openpyxl persists applyFont/applyFill (avoids white-on-white headers)."""
    if "verdictHeader" not in wb.named_styles:
        header = NamedStyle(name="verdictHeader")
        header.font = copy(HEADER_FONT)
        header.fill = copy(HEADER_FILL)
        header.alignment = copy(CENTER)
        header.border = copy(CELL_BORDER)
        wb.add_named_style(header)
    if "verdictDataLeft" not in wb.named_styles:
        left = NamedStyle(name="verdictDataLeft")
        left.font = copy(DATA_FONT)
        left.fill = copy(DATA_FILL)
        left.alignment = copy(LEFT)
        left.border = copy(CELL_BORDER)
        wb.add_named_style(left)
    if "verdictDataCenter" not in wb.named_styles:
        center = NamedStyle(name="verdictDataCenter")
        center.font = copy(DATA_FONT)
        center.fill = copy(DATA_FILL)
        center.alignment = copy(CENTER)
        center.border = copy(CELL_BORDER)
        wb.add_named_style(center)
    if "verdictTotal" not in wb.named_styles:
        total = NamedStyle(name="verdictTotal")
        total.font = copy(TOTAL_FONT)
        total.fill = copy(TOTAL_FILL)
        total.alignment = copy(CENTER)
        total.border = copy(CELL_BORDER)
        wb.add_named_style(total)


def _apply_header_style(cell) -> None:
    cell.style = "verdictHeader"


def _apply_data_style(cell, *, center: bool = True) -> None:
    cell.style = "verdictDataCenter" if center else "verdictDataLeft"


def _apply_total_style(cell) -> None:
    cell.style = "verdictTotal"


def _mechanic_value(breakdown: Dict[str, Any], key: str) -> int:
    """Mechanic counts always write 0 (never blank)."""
    return _safe_int((breakdown or {}).get(key), 0)


def _reason_value(text: Any) -> str | None:
    s = str(text or "").strip()
    return s if s else None


def _widths_from_template() -> Dict[str, float]:
    widths = dict(COLUMN_WIDTHS)
    if not TEMPLATE_PATH.is_file():
        return widths
    try:
        twb = load_workbook(TEMPLATE_PATH)
        tws = twb.active
        for col in range(1, tws.max_column + 1):
            letter = get_column_letter(col)
            dim = tws.column_dimensions.get(letter)
            if not (dim and dim.width):
                continue
            if col <= 3:
                widths[letter] = dim.width
            elif col == tws.max_column:
                widths[get_column_letter(COL_TOTAL)] = dim.width
            elif col == tws.max_column - 1:
                widths[get_column_letter(COL_ADDITIONAL_REASON)] = dim.width
            elif col == tws.max_column - 2:
                widths[get_column_letter(COL_ADDITIONAL)] = dim.width
            elif col == tws.max_column - 3:
                widths[get_column_letter(COL_APPEAL_REASON)] = dim.width
            elif col == tws.max_column - 4:
                widths[get_column_letter(COL_APPEAL)] = dim.width
    except Exception:
        pass
    return widths


def _force_style_apply_flags(path: Path) -> None:
    """Ensure cellXfs entries expose applyFont/applyFill/applyBorder for Excel/WPS."""
    import re
    import zipfile
    from io import BytesIO

    buf = BytesIO()
    with zipfile.ZipFile(path, "r") as zin:
        with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zout:
            for info in zin.infolist():
                data = zin.read(info.filename)
                if info.filename == "xl/styles.xml":
                    text = data.decode("utf-8")

                    def _patch_xf(match: re.Match[str]) -> str:
                        tag = match.group(0)
                        if "fontId=" not in tag:
                            return tag
                        for attr in ("applyFont", "applyFill", "applyBorder"):
                            if f'{attr}="' not in tag:
                                if tag.endswith("/>"):
                                    tag = tag[:-2] + f' {attr}="1"/>'
                                elif tag.endswith(">"):
                                    tag = tag[:-1] + f' {attr}="1">'
                        return tag

                    text = re.sub(r"<xf\b[^>]*/?>", _patch_xf, text)
                    data = text.encode("utf-8")
                # preserve original compress type for already-stored entries when possible
                zout.writestr(info, data)
    path.write_bytes(buf.getvalue())


def export_verdict_excel(
    payload: Dict[str, Any],
    target_dir: str | Path | None = None,
    boss_name: str = "宇宙之冕",
) -> Path:
    """
    生成终审 Excel（.xlsx）并写入磁盘。

    使用当前 HEADERS；次数为 0 写数字 0；仅「原因」「追加原因」允许空。
    样式按桌面样板配色显式写入（避免沿用模板 xf 索引导致表头样式丢失）。
    """
    date = str(payload.get("date") or datetime.now().date())
    players_raw: Sequence[Dict[str, Any]] = payload.get("players") or []
    points_per_count = _safe_int(payload.get("pointsPerCount"), 10)

    out_dir = resolve_export_dir(target_dir)
    out_path = _next_available_path(out_dir / f"智力表_{boss_name}_{date}.xlsx")

    wb = Workbook()
    ws = wb.active
    _ensure_named_styles(wb)
    ws.title = f"{boss_name}_{date.replace('-', '')}"[:31]
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30.0

    for col_letter, width in _widths_from_template().items():
        ws.column_dimensions[col_letter].width = width

    for col_idx, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        _apply_header_style(cell)

    for row_idx, player in enumerate(players_raw, start=2):
        breakdown = player.get("breakdown") or {}
        recognition = _safe_int(player.get("recognitionCount"), 0)
        appeal_count = _safe_int(player.get("appealAcquittalCount"), 0)
        additional_count = _safe_int(player.get("additionalCount"), 0)

        values: List[Any] = [
            str(player.get("name") or ""),
            str(player.get("rolesText") or ""),
            recognition,
            *[_mechanic_value(breakdown, key) for key, _ in MECHANIC_COLUMNS],
            appeal_count,
            _reason_value(player.get("appealAcquittalReasons")),
            additional_count,
            _reason_value(player.get("additionalReasons")),
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            center = col_idx != 1
            if col_idx in REASON_COLS:
                center = True
            _apply_data_style(cell, center=center)

        total_cell = ws.cell(
            row=row_idx,
            column=COL_TOTAL,
            value=(
                f"=({get_column_letter(COL_RECOGNITION)}{row_idx}"
                f"-{get_column_letter(COL_APPEAL)}{row_idx}"
                f"+{get_column_letter(COL_ADDITIONAL)}{row_idx})*{points_per_count}"
            ),
        )
        _apply_total_style(total_cell)

    wb.save(out_path)
    _force_style_apply_flags(out_path)
    return out_path
