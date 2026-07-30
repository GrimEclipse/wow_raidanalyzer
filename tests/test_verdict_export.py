import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from tools.export_verdict_excel import COL_TOTAL, export_verdict_excel


class VerdictExportTests(unittest.TestCase):
    def test_configured_points_per_count_is_used_in_excel_formula(self):
        payload = {
            "date": "2026-07-28",
            "pointsPerCount": 12,
            "players": [{
                "name": "Tester",
                "rolesText": "DPS",
                "recognitionCount": 3,
                "appealAcquittalCount": 1,
                "additionalCount": 0,
                "breakdown": {},
            }],
        }
        with tempfile.TemporaryDirectory() as directory:
            path = export_verdict_excel(payload, Path(directory), boss_name="Test")
            workbook = load_workbook(path, data_only=False)
            formula = workbook.active.cell(row=2, column=COL_TOTAL).value
        self.assertTrue(formula.endswith("*12"), formula)


if __name__ == "__main__":
    unittest.main()
